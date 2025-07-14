import os
import pandas as pd
import openpyxl
import re
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import glob
import ast
from utils import SafeFormulaEvaluator


def evaluate_weight_cell(pds_cell, sheet_values):
    """
    Evaluates the weight cell, handling both direct values and formulas.
    If the cell contains a formula, it safely evaluates it by substituting
    cell references with their actual values from the sheet.
    """
    if pds_cell.data_type == "f":  # It's a formula
        formula = pds_cell.value
        print(f"Evaluating formula: {formula}")
        expression = formula.lstrip("=")

        # Find all cell references (e.g., A1, $B$2)
        cell_references = re.findall(r"(\$?[A-Z]+\$?\d+)", expression)

        # Substitute references with their values from the values-only sheet
        for cell_ref in sorted(list(set(cell_references)), key=len, reverse=True):
            cell_value = sheet_values[cell_ref].value
            if cell_value is None:
                cell_value = 0
            expression = expression.replace(cell_ref, str(cell_value))

        # Safely evaluate the final expression
        try:
            evaluator = SafeFormulaEvaluator()
            node = ast.parse(expression, mode="eval").body
            result = evaluator.visit(node)
            print(f"Formula evaluated to: {result}")
            return result
        except (SyntaxError, NameError, TypeError, ValueError) as e:
            print(f"Could not evaluate formula {formula}: {e}")
            print("Falling back to formula string.")
            return pds_cell.value  # Fallback to the formula string on error
    else:  # It's a direct value
        return float(pds_cell.value)


def resolve_excel_references_in_sport_expression(
    sport_expression: str, sheet_values, pds_column_letter: str
) -> str:
    """
    Resolves Excel cell references within a sport expression string using values from the sheet.
    """
    if not isinstance(sport_expression, str):
        return ""

    expression = sport_expression.lstrip("=")

    # Find all cell references (e.g., A1, $B$2)
    cell_references = re.findall(r"(\$?[A-Z]+\$?\d+)", expression)

    # Substitute references with their values from the values-only sheet
    for cell_ref in sorted(list(set(cell_references)), key=len, reverse=True):
        # Extract the column part from cell_ref (e.g., "A" from "A1" or "$A$1")
        match = re.match(r"(\$?[A-Z]+)", cell_ref)
        if match:
            column_part = match.group(1).replace("$", "")
            if column_part == pds_column_letter:
                expression = expression.replace(cell_ref, "WEIGHT")
                continue  # Skip numerical substitution for weight references

        cell_value = sheet_values[cell_ref].value
        if cell_value is None:
            cell_value = 0
        expression = expression.replace(cell_ref, str(cell_value))

    return expression


def process_nutrition_journal():
    """
    Processes a nutrition journal from an Excel file.
    1. Finds the first .xlsx file in the current directory.
    2. Processes the "Journal" and "Variables" sheets.
    3. In the "Journal" sheet, it reconstructs dates by incrementing from the first valid date.
    4. Filters the "Journal" data for dates on or after 2024-06-30.
    5. Saves the processed "Journal" data to 'processed_journal.csv'.
    6. Saves the "Variables" data to 'variables.csv'.
    """
    # --- 1. Find the first .xlsx file ---
    data_dir = "data"
    xlsx_files = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if not xlsx_files:
        print(f"Error: No .xlsx file found in the '{data_dir}' directory.")
        return
    source_file = xlsx_files[0]

    # --- 2. Process "Journal" sheet ---
    workbook_formulas = openpyxl.load_workbook(source_file, data_only=False)
    sheet_formulas = workbook_formulas["Journal"]
    workbook_values = openpyxl.load_workbook(source_file, data_only=True)
    sheet_values = workbook_values["Journal"]

    header = [cell.value for cell in sheet_formulas[1]]
    date_col_idx = header.index("Date")
    pds_col_idx = header.index("Pds")
    nourriture_col_idx = header.index("Nourriture")
    sport_col_idx = header.index("Sport")

    pds_column_letter = get_column_letter(pds_col_idx + 1)

    data = []
    current_date = None
    first_date_found = False

    # --- 3. Reconstruct dates and combine data ---
    rows_formulas = sheet_formulas.iter_rows(min_row=2, max_row=sheet_formulas.max_row)
    rows_values = sheet_values.iter_rows(min_row=2, max_row=sheet_values.max_row)

    for row_formulas, row_values in zip(rows_formulas, rows_values):
        if not first_date_found:
            date_cell_value = row_formulas[date_col_idx].value
            if date_cell_value is not None and isinstance(date_cell_value, datetime):
                current_date = date_cell_value.date()
                first_date_found = True
        elif current_date:
            current_date += timedelta(days=1)
        pds_cell = row_formulas[pds_col_idx]
        if current_date < pd.to_datetime("2024-06-30").date() or pds_cell.value is None:
            continue
        pds_value = evaluate_weight_cell(pds_cell, sheet_values)
        data.append(
            {
                "Date": current_date,
                "Pds": pds_value,
                "Nourriture": row_formulas[nourriture_col_idx].value,
                "Sport": resolve_excel_references_in_sport_expression(
                    row_formulas[sport_col_idx].value,
                    sheet_values,
                    pds_column_letter,
                ),
            }
        )

    journal_df = pd.DataFrame(data)
    journal_df.dropna(subset=["Date"], inplace=True)
    journal_df["Date"] = pd.to_datetime(journal_df["Date"])

    # --- 4. Filter "Journal" data ---
    filtered_journal_df = journal_df[journal_df["Date"] >= "2024-06-30"].copy()

    # --- 5. Save processed "Journal" data ---
    output_journal_path = os.path.join(data_dir, "processed_journal.csv")
    filtered_journal_df.to_csv(output_journal_path, index=False)
    print(f"Successfully processed 'Journal' sheet and saved to {output_journal_path}")

    # --- 6. Process and save "Variables" sheet ---
    try:
        variables_df = pd.read_excel(source_file, sheet_name="Variables")
        output_variables_path = os.path.join(data_dir, "variables.csv")
        variables_df.to_csv(output_variables_path, index=False)
        print(
            f"Successfully processed 'Variables' sheet and saved to {output_variables_path}"
        )
    except Exception as e:
        print(f"Could not process 'Variables' sheet: {e}")


if __name__ == "__main__":
    process_nutrition_journal()
